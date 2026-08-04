"""Seed-set import: CSV/XLSX → validated ledger rows + a bad-row report.

Import-first and tolerant by design (docs/LEDGER-SCHEMA.md): single-number estimates are
accepted (stored as `estimate_single`), unknown modules go to a review queue instead of
rejecting the row, and every rejected row carries its reason — silent data loss poisons
calibration. One bad row must never abort or roll back the others: row-level failures
(including database-side rejections) land in the report via per-row savepoints.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.exc import DBAPIError
from sqlalchemy.ext.asyncio import AsyncSession

from estimo_core import LedgerEntry, ThreePoint, tr_lower
from estimo_knowledge.db import LedgerEntryRow

# Canonical column names (docs/LEDGER-SCHEMA.md import contract) + accepted aliases.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "brd_ref": ("brd_ref", "brd", "brd no", "doküman no", "dokuman no"),
    "brd_title": ("brd_title", "brd adı", "brd adi", "başlık", "baslik"),
    "customer": ("customer", "müşteri", "musteri"),
    "item_title": ("item_title", "kalem", "iş kalemi", "is kalemi", "kalem adı", "kalem adi"),
    "item_desc": ("item_desc", "item_description", "açıklama", "aciklama"),
    "modules": ("modules", "modül", "modul", "modüller", "moduller"),
    "domain": ("domain", "alan"),
    "team": ("team", "takım", "takim", "ekip"),
    "est_opt": ("est_opt", "iyimser"),
    "est_likely": ("est_likely", "olası", "olasi", "efor", "tahmin"),
    "est_pess": ("est_pess", "kötümser", "kotumser"),
    "est_single": ("est_single", "tek_efor"),
    "est_date": ("est_date", "tahmin tarihi"),
    "est_method": ("est_method", "yöntem", "yontem"),
    "actual_effort": ("actual_effort", "gerçekleşen", "gerceklesen"),
    "actual_source": ("actual_source", "gerçekleşme kaynağı", "gerceklesme kaynagi"),
    "completed_date": ("completed_date", "bitiş tarihi", "bitis tarihi"),
    "scope_changed": ("scope_changed", "kapsam değişti", "kapsam degisti"),
}

_TRUTHY = {"1", "true", "evet", "yes", "x"}
_FALSY = {"", "0", "false", "hayır", "hayir", "no"}

_DATE_FIELDS = ("est_date", "completed_date")
_FLOAT_FIELDS = ("est_opt", "est_likely", "est_pess", "est_single", "actual_effort")


@dataclass
class ImportReport:
    source: str
    total_rows: int = 0
    imported: int = 0
    rejected: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[dict[str, Any]] = field(default_factory=list)
    unknown_modules: dict[str, int] = field(default_factory=dict)
    # Rows that imported cleanly but carry no actual effort. They are NOT rejected —
    # an estimate whose actual has not landed yet is still a true record, and
    # LEDGER-SCHEMA.md's import contract is deliberately tolerant. They are counted
    # separately because they cannot answer the question this ledger exists for, so
    # calibration never sees them and the importing operator should know how many
    # of their rows are in that state.
    without_actuals: int = 0
    # Rows already present in the ledger, skipped rather than appended. A duplicate is
    # the same observation counted twice, and both readers of this table (calibration
    # and analog retrieval) treat every row as an independent sample.
    duplicates: list[dict[str, Any]] = field(default_factory=list)

    def summary(self) -> str:
        lines = [
            f"seed import: {self.source}",
            (
                f"  rows: {self.total_rows}  imported: {self.imported}  "
                f"rejected: {len(self.rejected)}  warnings: {len(self.warnings)}"
            ),
        ]
        if self.without_actuals:
            lines.append(f"  imported without actuals (not in calibration): {self.without_actuals}")
        if self.duplicates:
            lines.append(f"  skipped as already in the ledger: {len(self.duplicates)}")
        if self.unknown_modules:
            listed = ", ".join(f"{m}×{c}" for m, c in sorted(self.unknown_modules.items()))
            lines.append(f"  review queue (unknown modules): {listed}")
        for row in self.rejected[:20]:
            lines.append(f"  row {row['row']}: {row['error']}")
        for row in self.warnings[:20]:
            lines.append(f"  row {row['row']} (warning): {row['warning']}")
        return "\n".join(lines)


def _read_csv(path: Path) -> list[dict[str, str]]:
    raw = path.read_text(encoding="utf-8-sig")
    if not raw.strip():
        return []
    header_line = raw.splitlines()[0]
    # Sniff on parsed cell counts so delimiters inside quoted cells can't flip the dialect.
    counts = {d: len(next(csv.reader([header_line], delimiter=d), [])) for d in (";", ",")}
    delimiter = ";" if counts[";"] > counts[","] else ","
    reader = csv.DictReader(io.StringIO(raw, newline=""), delimiter=delimiter)
    return [dict(row) for row in reader]


def _cell_str(cell: object) -> str:
    if isinstance(cell, dt.datetime):
        return cell.date().isoformat()
    if isinstance(cell, dt.date):
        return cell.isoformat()
    return "" if cell is None else str(cell)


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return []
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [_cell_str(cell).strip() for cell in next(rows)]
    except StopIteration:
        return []
    return [{header[i]: _cell_str(cell) for i, cell in enumerate(row)} for row in rows]


def read_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    return _read_csv(path)


def propose_mapping(headers: list[str]) -> dict[str, str | None]:
    """Header → canonical field, by alias, for the import wizard's mapping step.

    Returns an entry for EVERY header, `None` where nothing matched, so the wizard
    can show "Column G → unmapped" instead of quietly dropping it. A field already
    claimed by an earlier header is not claimed twice — two columns feeding one
    field would make which one wins depend on dictionary order.
    """
    proposal: dict[str, str | None] = {}
    claimed: set[str] = set()
    by_alias = {alias: name for name, aliases in COLUMN_ALIASES.items() for alias in aliases}
    for header in headers:
        stripped = (header or "").strip()
        field_name = by_alias.get(tr_lower(stripped)) or by_alias.get(stripped.lower())
        if field_name is not None and field_name not in claimed:
            claimed.add(field_name)
            proposal[stripped] = field_name
        else:
            proposal[stripped] = None
    return proposal


def _canonicalize(row: dict[Any, Any], mapping: dict[str, str] | None = None) -> dict[str, str]:
    if None in row:
        msg = "row has more fields than the header (stray delimiter?)"
        raise ValueError(msg)
    if mapping is not None:
        # A confirmed mapping is the WHOLE contract: no alias fallback underneath it.
        # An operator who unmapped a column in the wizard must not find it re-mapped
        # by a name coincidence — that would import a column they chose to exclude.
        confirmed: dict[str, str] = {}
        for key, value in row.items():
            field_name = mapping.get((key or "").strip())
            val = value.strip() if isinstance(value, str) else ""
            if field_name and val != "":
                confirmed.setdefault(field_name, val)
        return confirmed
    canonical: dict[str, str] = {}
    lowered: dict[str, str] = {}
    for key, value in row.items():
        stripped = (key or "").strip()
        val = value.strip() if isinstance(value, str) else ""
        # Register under BOTH normalizations: tr_lower handles Turkish İ/I headers,
        # plain lower keeps ALL-CAPS English headers (ITEM_TITLE) intact.
        for norm in (tr_lower(stripped), stripped.lower()):
            if val != "":
                lowered.setdefault(norm, val)
    for field_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in lowered:
                canonical[field_name] = lowered[alias]
                break
    return canonical


def _parse_date(value: str) -> dt.date | None:
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(value, fmt).date()  # noqa: DTZ007
        except ValueError:
            continue
    return None


def _parse_float(value: str) -> float | None:
    cleaned = value.strip()
    if "," in cleaned:
        # Turkish notation: dots are thousands separators, comma is the decimal mark.
        cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _collect_parse_warnings(canonical: dict[str, str]) -> list[str]:
    """No-silent-loss contract: a non-empty value that fails to parse is reported."""
    warnings: list[str] = []
    for field_name in _DATE_FIELDS:
        if canonical.get(field_name) and _parse_date(canonical[field_name]) is None:
            warnings.append(f"{field_name}: unparseable date {canonical[field_name]!r}")
    for field_name in _FLOAT_FIELDS:
        if canonical.get(field_name) and _parse_float(canonical[field_name]) is None:
            warnings.append(f"{field_name}: unparseable number {canonical[field_name]!r}")
    return warnings


def to_ledger_entry(row: dict[str, str]) -> LedgerEntry:
    """Validate one canonicalized row through the domain model (raises on bad rows)."""
    opt = _parse_float(row.get("est_opt", "")) if row.get("est_opt") else None
    likely = _parse_float(row.get("est_likely", "")) if row.get("est_likely") else None
    pess = _parse_float(row.get("est_pess", "")) if row.get("est_pess") else None
    single = _parse_float(row.get("est_single", "")) if row.get("est_single") else None

    estimate: ThreePoint | None = None
    estimate_single: float | None = None
    if opt is not None and likely is not None and pess is not None:
        estimate = ThreePoint(optimistic=opt, likely=likely, pessimistic=pess)
    elif single is not None:
        estimate_single = single
    elif likely is not None:
        estimate_single = likely

    scope_raw = tr_lower(row.get("scope_changed", ""))
    if scope_raw and scope_raw not in _TRUTHY | _FALSY:
        msg = f"scope_changed value not recognized: {scope_raw!r}"
        raise ValueError(msg)

    return LedgerEntry(
        brd_ref=row.get("brd_ref", ""),
        item_title=row.get("item_title", ""),
        item_description=row.get("item_desc") or None,
        module_tags=tuple(
            m.strip() for m in row.get("modules", "").replace(";", ",").split(",") if m.strip()
        ),
        # Slice keys are compared, so they are normalized at the boundary — the same
        # rule the product write applies (estimo_estimate.loop), or the seed set and
        # the product would put two vocabularies in one column.
        domain_tags=tuple(
            tag for tag in (tr_lower(d.strip()) for d in row.get("domain", "").split(",")) if tag
        ),
        team=tr_lower((row.get("team") or "").strip()) or None,
        estimate=estimate,
        estimate_single=estimate_single,
        estimated_at=_parse_date(row["est_date"]) if row.get("est_date") else None,
        method=row.get("est_method") or None,
        actual_effort=_parse_float(row["actual_effort"]) if row.get("actual_effort") else None,
        actual_source=row.get("actual_source") or None,
        completed_at=_parse_date(row["completed_date"]) if row.get("completed_date") else None,
        scope_changed=scope_raw in _TRUTHY,
    )


def _to_row(entry: LedgerEntry, extras: dict[str, str]) -> LedgerEntryRow:
    return LedgerEntryRow(
        brd_ref=entry.brd_ref,
        brd_title=extras.get("brd_title") or None,
        customer=extras.get("customer") or None,
        item_title=entry.item_title,
        item_description=entry.item_description,
        module_tags=list(entry.module_tags),
        domain_tags=list(entry.domain_tags),
        team=entry.team,
        est_optimistic=entry.estimate.optimistic if entry.estimate else None,
        est_likely=entry.estimate.likely if entry.estimate else None,
        est_pessimistic=entry.estimate.pessimistic if entry.estimate else None,
        estimate_single=entry.estimate_single,
        estimated_at=entry.estimated_at,
        method=entry.method,
        actual_effort=entry.actual_effort,
        actual_source=entry.actual_source,
        completed_at=entry.completed_at,
        scope_changed=entry.scope_changed,
    )


# What makes two ledger rows "the same job". Deliberately the whole observation, not
# just an id: two genuinely different jobs may share a BRD reference and a title (the
# same item re-estimated for another team), and those differ in team or in the numbers.
# A row matching on ALL of these is the same measurement, arriving twice.
def _dedupe_key(entry: LedgerEntry) -> tuple[Any, ...]:
    return (
        entry.brd_ref,
        entry.item_title,
        entry.team,
        entry.estimate.optimistic if entry.estimate else None,
        entry.estimate.likely if entry.estimate else None,
        entry.estimate.pessimistic if entry.estimate else None,
        entry.estimate_single,
        entry.actual_effort,
        entry.completed_at,
    )


async def _existing_keys(session: AsyncSession) -> set[tuple[Any, ...]]:
    """Dedupe keys already in the ledger, read once per import.

    Read up front rather than queried per row: an import is one pass over a file, and
    a per-row SELECT would turn a 10k-row seed set into 10k round-trips. Only rows the
    caller's tenant can see are returned — RLS applies to this SELECT like any other.
    """
    result = await session.execute(
        select(
            LedgerEntryRow.brd_ref,
            LedgerEntryRow.item_title,
            LedgerEntryRow.team,
            LedgerEntryRow.est_optimistic,
            LedgerEntryRow.est_likely,
            LedgerEntryRow.est_pessimistic,
            LedgerEntryRow.estimate_single,
            LedgerEntryRow.actual_effort,
            LedgerEntryRow.completed_at,
        )
    )
    return {
        (
            brd_ref,
            item_title,
            team,
            _as_float(opt),
            _as_float(likely),
            _as_float(pess),
            _as_float(single),
            _as_float(actual),
            completed,
        )
        for brd_ref, item_title, team, opt, likely, pess, single, actual, completed in result
    }


def _as_float(value: Any) -> float | None:
    """Numeric columns come back as Decimal; the entry side holds floats."""
    return None if value is None else float(value)


async def import_rows(
    session: AsyncSession,
    rows: list[dict[str, str]],
    *,
    source: str,
    taxonomy: set[str] | None = None,
    mapping: dict[str, str] | None = None,
) -> ImportReport:
    """Import already-parsed rows; commits once at the end.

    The file-reading half lives in `import_seed`; this half is what the API's
    upload wizard drives, so the CLI and the panel import through exactly the same
    validation, the same per-row savepoints and the same report.
    """
    report = ImportReport(source=source)
    seen = await _existing_keys(session)
    for row_no, raw in enumerate(rows, start=2):  # header is line 1
        report.total_rows += 1
        try:
            canonical = _canonicalize(raw, mapping)
            entry = to_ledger_entry(canonical)
        except (ValidationError, ValueError) as exc:
            message = str(exc)
            if isinstance(exc, ValidationError):
                first = exc.errors()[0]
                message = f"{'.'.join(str(loc) for loc in first['loc'])}: {first['msg']}"
            report.rejected.append({"row": row_no, "error": message})
            continue
        key = _dedupe_key(entry)
        if key in seen:
            # An identical row is already in the ledger. Importing it again is never a
            # second observation of the world — it is the same observation counted
            # twice, and the ledger's readers are calibration and analog retrieval, so
            # a duplicated row both inflates the sample it is graded against and
            # becomes its own nearest analog. Re-running an import (a retry, a file
            # sent twice, a wizard the operator was unsure had worked) must therefore
            # be a no-op for rows already present, and must SAY so.
            report.duplicates.append({"row": row_no, "brd_ref": entry.brd_ref})
            continue
        seen.add(key)
        for warning in _collect_parse_warnings(canonical):
            report.warnings.append({"row": row_no, "warning": warning})
        if taxonomy is not None:
            for module in entry.module_tags:
                if module not in taxonomy:
                    report.unknown_modules[module] = report.unknown_modules.get(module, 0) + 1
        try:
            # Per-row savepoint: a database-side rejection (e.g. varchar overflow)
            # must reject THIS row, not roll back the whole import at commit time.
            async with session.begin_nested():
                session.add(_to_row(entry, canonical))
                await session.flush()
        except DBAPIError as exc:
            report.rejected.append({"row": row_no, "error": f"db rejected row: {exc.orig}"})
            continue
        report.imported += 1
        if entry.actual_effort is None:
            report.without_actuals += 1
    await session.commit()
    return report


async def import_seed(
    session: AsyncSession,
    path: Path,
    *,
    taxonomy: set[str] | None = None,
    mapping: dict[str, str] | None = None,
) -> ImportReport:
    """Import a seed file; commits once at the end. Bad rows never abort good rows."""
    return await import_rows(
        session, read_rows(path), source=path.name, taxonomy=taxonomy, mapping=mapping
    )
